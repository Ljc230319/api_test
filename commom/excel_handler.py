import openpyxl
from config.setting import config


class ExcelHandler():
    def __init__(self,file):
        self.file = file

     #打开表单
    def open_sheet(self,sheet_name):
        wb = openpyxl.load_workbook(self.file)
        sheet = wb[sheet_name]
        return sheet

    #读取表单第一行(表头)
    def header(self,sheet_name):
        sheet = self.open_sheet(sheet_name)  #调用openexcel函数获取表单
        headers = []
        for i in sheet[1]:
            headers.append(i.value)
        return headers  #获取第一行

    # 读取所有的数据
    def read(self,sheet_name):
        sheet = self.open_sheet(sheet_name)
        rows = list(sheet.rows)
        data = []
        for row in rows[1:]:
            row_data = []
            for cell in row:
                print(cell.value)
                row_data.append(cell.value)
            data_dict = dict(zip(self.header(sheet_name),row_data))
            data.append(data_dict)
        return data
    @staticmethod
    def write(file,sheet_name,row,column,data):
        wb = openpyxl.load_workbook(file)
        sheet = wb[sheet_name]
        sheet.cell(row,column).value = data   #修改单元格
        wb.save(file)    #保存
        wb.close()       #关闭

if __name__ == '__main__':
    excel = ExcelHandler('/Users/apple/Desktop/case.xlsx')
    header = excel.header('Sheet3')
    data = excel.read('Sheet3')
    print(data)










